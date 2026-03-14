from __future__ import annotations

import certifi
import gzip
import io
import json
import math
import random
import re
import statistics
import zipfile
from collections import Counter, defaultdict
from datetime import UTC, datetime, timedelta
from pathlib import Path

import openpyxl
import requests


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
RAW_DIR = DATA_DIR / "raw"
CURATED_DIR = DATA_DIR / "curated"

SEED = 20260314
RNG = random.Random(SEED)

AMAZON_SOURCES = [
    {
        "name": "amazon_industrial_scientific",
        "dataset": "Amazon Reviews 2023 Industrial and Scientific",
        "url": "https://mcauleylab.ucsd.edu/public_datasets/data/amazon_2023/raw/meta_categories/meta_Industrial_and_Scientific.jsonl.gz",
        "candidate_limit": 220,
    },
    {
        "name": "amazon_tools_home_improvement",
        "dataset": "Amazon Reviews 2023 Tools and Home Improvement",
        "url": "https://mcauleylab.ucsd.edu/public_datasets/data/amazon_2023/raw/meta_categories/meta_Tools_and_Home_Improvement.jsonl.gz",
        "candidate_limit": 220,
    },
]

UCI_SOURCE = {
    "name": "uci_online_retail_ii",
    "dataset": "UCI Online Retail II",
    "url": "https://archive.ics.uci.edu/static/public/502/online%2Bretail%2Bii.zip",
}

PRODUCT_TARGETS = {
    "Raw Materials": 32,
    "Finished Goods": 20,
    "Electronics": 30,
    "Packaging": 26,
    "Tools & Equipment": 32,
}

CATEGORY_DEFINITIONS = [
    {"name": "Raw Materials", "description": "Curated industrial consumables, adhesives, tubing, stock material, and shop-floor inputs."},
    {"name": "Finished Goods", "description": "Packaged finished items and assembled inventory ready for dispatch."},
    {"name": "Electronics", "description": "Electronic components, control hardware, wiring accessories, and powered modules."},
    {"name": "Packaging", "description": "Packaging and storage supplies used for handling, labeling, and outbound preparation."},
    {"name": "Tools & Equipment", "description": "Tools, safety gear, fixtures, and workshop equipment."},
]

CATEGORY_PREFIX = {
    "Raw Materials": "RM",
    "Finished Goods": "FG",
    "Electronics": "EL",
    "Packaging": "PK",
    "Tools & Equipment": "TE",
}

CATEGORY_LOCATIONS = {
    "Raw Materials": [("WH1/Main_Store", 0.50), ("WH2/Storage", 0.25), ("WH3/Production", 0.25)],
    "Finished Goods": [("WH3/Production", 0.60), ("WH1/Shipping", 0.40)],
    "Electronics": [("WH1/Rack_A", 0.45), ("WH1/Main_Store", 0.25), ("WH3/QC_Area", 0.30)],
    "Packaging": [("WH1/Rack_B", 0.45), ("WH3/Production", 0.35), ("WH1/Receiving", 0.20)],
    "Tools & Equipment": [("WH2/Storage", 0.55), ("WH3/QC_Area", 0.25), ("WH1/Main_Store", 0.20)],
}

SOURCE_TO_INTERNAL_CATEGORY = {
    "Amazon Reviews 2023 Industrial and Scientific": "Industrial & Scientific",
    "Amazon Reviews 2023 Tools and Home Improvement": "Tools & Home Improvement",
}


def ensure_dirs() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    CURATED_DIR.mkdir(parents=True, exist_ok=True)


def session() -> requests.Session:
    http = requests.Session()
    http.trust_env = False
    http.verify = certifi.where()
    http.headers.update({"User-Agent": "CoreInventory-Curator/1.0"})
    return http


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def clean_title(title: str) -> str:
    normalized = re.sub(r"\s+", " ", title).strip()
    normalized = normalized.replace("Protect Your Warehouse & Business", "").strip(" -|,")
    if len(normalized) > 86:
        normalized = normalized[:83].rstrip(" ,-/") + "..."
    return normalized


def looks_valid_title(title: str) -> bool:
    if not title or len(title) < 12:
        return False
    junk = ["amazon", "ships from", "visit the", "free shipping"]
    consumer = [
        "christmas", "halloween", "wedding", "birthday", "bedside", "nightstand",
        "chandelier", "nursery", "party", "living room", "bedroom", "bathroom",
        "valentine", "holiday", "home decor", "sofa", "garden decor", "vanity",
    ]
    lowered = title.lower()
    return not any(token in lowered for token in junk + consumer)


def extract_brand(record: dict) -> str | None:
    details = record.get("details") or {}
    for key in ("Brand", "Manufacturer"):
        value = details.get(key)
        if value:
            return str(value).strip()
    store = record.get("store")
    return str(store).strip() if store else None


def category_text(record: dict) -> str:
    pieces = []
    for key in ("main_category", "title", "store"):
        value = record.get(key)
        if value:
            pieces.append(str(value))
    for value in record.get("categories") or []:
        pieces.append(str(value))
    return " ".join(pieces).lower()


def classify_category(record: dict) -> str:
    text = category_text(record)

    electronics_keywords = [
        "sensor", "electrical", "electronic", "connector", "switch", "led", "cable",
        "battery", "charger", "adapter", "module", "circuit", "voltage",
        "usb", "soldering", "relay", "wiring",
    ]
    packaging_keywords = [
        "tape", "wrap", "label", "mailer", "box", "bag", "bubble", "shrink",
        "stretch", "container", "lid", "storage bin", "pouch", "tray", "carton",
        "packing", "shipping", "barcode",
    ]
    tools_keywords = [
        "wrench", "drill", "screwdriver", "pliers", "clamp", "hammer", "tool",
        "glove", "safety", "helmet", "goggles", "meter", "torch", "ratchet",
        "socket", "cutter", "knife", "saw", "ladder", "trolley",
    ]
    raw_keywords = [
        "steel", "aluminum", "copper", "brass", "tubing", "hose", "sheet",
        "rod", "adhesive", "sealant", "resin", "lubricant", "solvent", "powder",
        "foam", "rubber", "gasket", "wire rope", "welding", "metal",
    ]
    finished_keywords = [
        "dispenser", "cabinet", "organizer", "holder", "rack", "sign", "assembly",
        "fixture", "case", "frame", "station", "unit", "kit",
    ]

    if any(keyword in text for keyword in raw_keywords):
        return "Raw Materials"
    if any(keyword in text for keyword in packaging_keywords):
        return "Packaging"
    if any(keyword in text for keyword in electronics_keywords):
        return "Electronics"
    if any(keyword in text for keyword in finished_keywords):
        return "Finished Goods"
    if any(keyword in text for keyword in tools_keywords):
        return "Tools & Equipment"

    if "industrial & scientific" in text:
        return "Raw Materials"
    if "tools & home improvement" in text:
        return "Tools & Equipment"
    return "Finished Goods"


def infer_uom(record: dict, category: str) -> str:
    text = category_text(record)
    if any(token in text for token in ("roll", "tape", "wrap", "label")):
        return "rolls"
    if any(token in text for token in ("cable", "wire", "tubing", "hose")):
        return "m"
    if any(token in text for token in ("glove", "goggle", "shoe", "sock")):
        return "pairs"
    if any(token in text for token in ("kit", "set", "assortment")):
        return "sets"
    if any(token in text for token in ("powder", "sealant", "adhesive", "lubricant", "resin", "paint")):
        return "kg"
    if category == "Raw Materials" and any(token in text for token in ("film", "wire", "sheet")):
        return "m"
    return "pcs"


def quality_score(record: dict) -> float:
    rating_number = float(record.get("rating_number") or 0)
    average_rating = float(record.get("average_rating") or 0)
    price = float(record.get("price") or 0)
    return math.log1p(rating_number) * 15 + average_rating * 8 + min(price, 200) / 20


def stream_amazon_candidates(http: requests.Session, source: dict) -> tuple[list[dict], int]:
    response = http.get(source["url"], stream=True, timeout=180)
    response.raise_for_status()

    seen = set()
    candidates = []
    scanned = 0

    with gzip.GzipFile(fileobj=response.raw) as gz:
        for raw_line in gz:
            scanned += 1
            record = json.loads(raw_line.decode("utf-8"))
            title = clean_title(str(record.get("title") or ""))
            if not looks_valid_title(title):
                continue
            if record.get("price") in (None, "", "None"):
                continue
            key = slug(title)
            if key in seen:
                continue

            category = classify_category(record)
            brand = extract_brand(record)
            candidate = {
                "name": title,
                "normalized_name": key,
                "brand": brand,
                "source_dataset": source["dataset"],
                "source_url": source["url"],
                "source_ref": record.get("parent_asin"),
                "price": float(record.get("price") or 0),
                "rating_number": int(record.get("rating_number") or 0),
                "average_rating": float(record.get("average_rating") or 0),
                "raw_categories": record.get("categories") or [],
                "category": category,
                "uom": infer_uom(record, category),
                "score": quality_score(record),
            }
            candidates.append(candidate)
            seen.add(key)

            if len(candidates) >= source["candidate_limit"]:
                break

    return candidates, scanned


def load_uci_profile(http: requests.Session) -> dict:
    response = http.get(UCI_SOURCE["url"], timeout=240)
    response.raise_for_status()

    archive = zipfile.ZipFile(io.BytesIO(response.content))
    workbook_path = RAW_DIR / "online_retail_ii.xlsx"
    workbook_path.write_bytes(archive.read("online_retail_II.xlsx"))

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    positive_quantities: list[int] = []
    line_values: list[float] = []
    invoice_counts: Counter[str] = Counter()
    countries: Counter[str] = Counter()
    rows_processed = 0

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                invoice, stock_code, description, quantity, invoice_date, price, customer_id, country = row
                rows_processed += 1

                if not quantity or quantity <= 0 or not price or price <= 0 or not description:
                    continue

                quantity = int(quantity)
                price = float(price)
                positive_quantities.append(min(quantity, 240))
                line_values.append(min(quantity * price, 5000))

                if invoice_date:
                    invoice_counts[str(invoice_date.date())] += 1
                if country:
                    countries[str(country)] += 1

                if len(positive_quantities) >= 180000:
                    break
            if len(positive_quantities) >= 180000:
                break
    finally:
        workbook.close()

    positive_quantities.sort()
    line_values.sort()
    avg_invoices_per_day = round(sum(invoice_counts.values()) / max(len(invoice_counts), 1), 2)

    profile = {
        "dataset": UCI_SOURCE["dataset"],
        "source_url": UCI_SOURCE["url"],
        "rows_processed": rows_processed,
        "quantity_sample_size": len(positive_quantities),
        "quantity_quantiles": {
            "p25": positive_quantities[int(len(positive_quantities) * 0.25)],
            "p50": positive_quantities[int(len(positive_quantities) * 0.50)],
            "p75": positive_quantities[int(len(positive_quantities) * 0.75)],
            "p90": positive_quantities[int(len(positive_quantities) * 0.90)],
        },
        "value_quantiles": {
            "p25": round(line_values[int(len(line_values) * 0.25)], 2),
            "p50": round(line_values[int(len(line_values) * 0.50)], 2),
            "p75": round(line_values[int(len(line_values) * 0.75)], 2),
            "p90": round(line_values[int(len(line_values) * 0.90)], 2),
        },
        "avg_invoice_lines_per_day": avg_invoices_per_day,
        "top_countries": countries.most_common(8),
        "quantities": positive_quantities,
    }

    summary_path = RAW_DIR / "uci_online_retail_ii_profile.json"
    summary_path.write_text(json.dumps({k: v for k, v in profile.items() if k != "quantities"}, indent=2), encoding="utf-8")

    return profile


def select_products(candidates: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for candidate in candidates:
        grouped[candidate["category"]].append(candidate)

    for items in grouped.values():
        items.sort(key=lambda item: item["score"], reverse=True)

    selected: list[dict] = []
    used = set()

    for category, target in PRODUCT_TARGETS.items():
        picked = 0
        for candidate in grouped.get(category, []):
            if candidate["normalized_name"] in used:
                continue
            selected.append(candidate)
            used.add(candidate["normalized_name"])
            picked += 1
            if picked >= target:
                break

    if len(selected) < sum(PRODUCT_TARGETS.values()):
        remaining = sorted(
            [candidate for candidate in candidates if candidate["normalized_name"] not in used],
            key=lambda item: item["score"],
            reverse=True,
        )
        for candidate in remaining:
            selected.append(candidate)
            used.add(candidate["normalized_name"])
            if len(selected) >= sum(PRODUCT_TARGETS.values()):
                break

    selected.sort(key=lambda item: (item["category"], -item["score"], item["name"]))
    return selected


def demand_signal_for(candidate: dict, quantity_profile: dict) -> float:
    rating_factor = math.log1p(candidate["rating_number"] + 1)
    price_factor = 0.8 if candidate["price"] >= 120 else 1.0
    popularity = max(0.7, 0.9 + rating_factor / 3)
    base_qty = RNG.choice(quantity_profile["quantities"])
    return base_qty * popularity * price_factor


def reorder_bounds(candidate: dict, quantity_profile: dict) -> tuple[int, int]:
    demand_signal = demand_signal_for(candidate, quantity_profile)
    uom_floor = {
        "kg": 18,
        "m": 60,
        "pcs": 20,
        "rolls": 8,
        "sets": 5,
        "pairs": 12,
    }
    category_multiplier = {
        "Raw Materials": 1.10,
        "Finished Goods": 0.90,
        "Electronics": 0.80,
        "Packaging": 1.20,
        "Tools & Equipment": 0.70,
    }
    reorder_min = int(max(uom_floor.get(candidate["uom"], 10), demand_signal * category_multiplier[candidate["category"]] / 3.2))
    reorder_max = int(max(reorder_min + 5, reorder_min * (4.2 if candidate["category"] != "Tools & Equipment" else 3.4)))
    return reorder_min, reorder_max


def allocate_stock(category: str, total_stock: int) -> list[dict]:
    locations = CATEGORY_LOCATIONS[category]
    stock_rows = []
    remaining = total_stock
    for index, (location, share) in enumerate(locations):
        if index == len(locations) - 1:
            qty = remaining
        else:
            qty = max(1, int(round(total_stock * share)))
            remaining -= qty
        if qty > 0:
            stock_rows.append({"location": location, "qty": qty})
    return stock_rows


def build_products(selected: list[dict], quantity_profile: dict) -> list[dict]:
    products = []
    counters = Counter()

    for candidate in selected:
        counters[candidate["category"]] += 1
        reorder_min, reorder_max = reorder_bounds(candidate, quantity_profile)
        if counters[candidate["category"]] % 9 == 0:
            total_stock = max(1, int(reorder_min * 0.75))
        else:
            total_stock = int(reorder_min * (2.5 + (candidate["score"] % 3)))
        total_stock = min(total_stock, reorder_max * 2)
        sku = f"{CATEGORY_PREFIX[candidate['category']]}-{counters[candidate['category']]:04d}"

        products.append(
            {
                "name": candidate["name"],
                "sku": sku,
                "category": candidate["category"],
                "uom": candidate["uom"],
                "reorder_min": reorder_min,
                "reorder_max": reorder_max,
                "stock_by_location": allocate_stock(candidate["category"], total_stock),
                "source_dataset": candidate["source_dataset"],
                "source_ref": candidate["source_ref"],
                "brand": candidate["brand"],
                "unit_cost": round(candidate["price"], 2),
                "rating_number": candidate["rating_number"],
                "average_rating": candidate["average_rating"],
            }
        )

    return sorted(products, key=lambda item: (item["category"], item["name"]))


def weighted_product_choice(products: list[dict]) -> dict:
    weights = [max(1.0, math.log1p(product.get("rating_number", 0) + 1) * 4 + product["reorder_min"] / 12) for product in products]
    return RNG.choices(products, weights=weights, k=1)[0]


def scale_quantity(base_qty: int, uom: str, operation_type: str) -> int:
    qty = base_qty
    if uom == "kg":
        qty = max(3, round(base_qty / 2))
    elif uom == "m":
        qty = max(8, round(base_qty * 1.8))
    elif uom == "rolls":
        qty = max(1, round(base_qty / 5))
    elif uom == "sets":
        qty = max(1, round(base_qty / 8))
    elif uom == "pairs":
        qty = max(1, round(base_qty / 3))

    if operation_type == "adjustment":
        qty = max(1, round(qty * 0.15))
    elif operation_type == "transfer":
        qty = max(1, round(qty * 0.7))

    return int(min(qty, 240))


def primary_location_for(product: dict) -> str:
    return product["stock_by_location"][0]["location"]


def delivery_location_for(product: dict) -> str:
    locations = [row["location"] for row in product["stock_by_location"]]
    if product["category"] == "Finished Goods" and "WH1/Shipping" in locations:
        return "WH1/Shipping"
    return locations[0]


def transfer_destination(product: dict) -> str:
    if product["category"] == "Raw Materials":
        return "WH3/Production"
    if product["category"] == "Packaging":
        return "WH3/Production"
    if product["category"] == "Electronics":
        return "WH3/QC_Area"
    if product["category"] == "Tools & Equipment":
        return "WH2/Storage"
    return "WH1/Shipping"


def generate_history(products: list[dict], quantity_profile: dict) -> list[dict]:
    counters = Counter()
    history = []
    manager = "manager@coreinventory.com"
    staff = "staff@coreinventory.com"
    operation_weights = [("receipt", 0.28), ("delivery", 0.34), ("transfer", 0.25), ("adjustment", 0.13)]
    operation_names = [name for name, _ in operation_weights]
    weights = [weight for _, weight in operation_weights]
    today = datetime.now(UTC).replace(hour=8, minute=0, second=0, microsecond=0)

    for days_ago in range(44, -1, -1):
        move_count = 6 + int(RNG.random() * 6)
        day_base = today - timedelta(days=days_ago)

        for move_index in range(move_count):
            product = weighted_product_choice(products)
            operation = RNG.choices(operation_names, weights=weights, k=1)[0]
            counters[operation] += 1
            qty = scale_quantity(RNG.choice(quantity_profile["quantities"]), product["uom"], operation)
            created_at = (day_base + timedelta(minutes=move_index * 37)).isoformat().replace("+00:00", "Z")
            from_location = primary_location_for(product)
            to_location = primary_location_for(product)

            if operation == "receipt":
                from_location = "VENDOR"
                to_location = primary_location_for(product)
            elif operation == "delivery":
                from_location = delivery_location_for(product)
                to_location = "CUSTOMER"
            elif operation == "transfer":
                from_location = primary_location_for(product)
                to_location = transfer_destination(product)
                if to_location == from_location:
                    to_location = "WH1/Shipping"
            else:
                is_positive = RNG.random() > 0.65
                from_location = "ADJUSTMENT" if is_positive else primary_location_for(product)
                to_location = primary_location_for(product) if is_positive else "ADJUSTMENT"

            history.append(
                {
                    "operation_type": operation,
                    "ref": f"{operation[:3].upper()}-EXT-{counters[operation]:05d}",
                    "sku": product["sku"],
                    "from_location": from_location,
                    "to_location": to_location,
                    "qty": qty,
                    "user_email": staff if operation == "delivery" else manager,
                    "created_at": created_at,
                    "source_dataset": UCI_SOURCE["dataset"] if operation in {"receipt", "delivery"} else "Derived warehouse operations",
                }
            )

    history.sort(key=lambda row: row["created_at"])
    return history


def group_products(products: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for product in products:
        grouped[product["category"]].append(product)
    return grouped


def top_products(products: list[dict], count: int, category: str | None = None) -> list[dict]:
    filtered = [product for product in products if category is None or product["category"] == category]
    return sorted(filtered, key=lambda item: (item["reorder_max"], item.get("rating_number", 0)), reverse=True)[:count]


def generate_operations(products: list[dict], quantity_profile: dict) -> dict:
    grouped = group_products(products)
    manager = "manager@coreinventory.com"
    staff = "staff@coreinventory.com"

    suppliers = [
        "Prime Industrial Supply",
        "Northline Components",
        "Applied Shop Materials",
        "Vector Warehouse Products",
        "Meridian Tools Distribution",
        "Axis Packaging & Storage",
    ]
    customers = [
        "BuildCorp Ltd.",
        "Northwind Fabrication",
        "Vertex Distribution",
        "Apex Maintenance Services",
        "Cobalt Manufacturing",
        "SignalWorks Integrators",
    ]

    receipts = []
    receipt_products = top_products(products, 18)
    for index in range(0, len(receipt_products), 3):
        lines = []
        for product in receipt_products[index:index + 3]:
            location = {
                "Raw Materials": "WH1/Receiving",
                "Electronics": "WH1/Rack_A",
                "Packaging": "WH1/Receiving",
                "Tools & Equipment": "WH2/Storage",
                "Finished Goods": "WH3/Production",
            }[product["category"]]
            lines.append(
                {
                    "sku": product["sku"],
                    "expected_qty": scale_quantity(RNG.choice(quantity_profile["quantities"]), product["uom"], "receipt"),
                    "received_qty": 0,
                    "location": location,
                }
            )
        receipts.append(
            {
                "ref": f"REC-{index // 3 + 1:05d}",
                "supplier": suppliers[(index // 3) % len(suppliers)],
                "warehouse": "WH1" if index % 2 == 0 else "WH2",
                "scheduled_in_days": (index // 3) % 6,
                "status": ["draft", "waiting", "ready"][index % 3],
                "created_by": manager if index % 2 == 0 else staff,
                "notes": "Curated from external catalog demand profile.",
                "lines": lines,
            }
        )

    deliveries = []
    delivery_products = top_products(products, 18, None)
    for index in range(0, len(delivery_products), 2):
        lines = []
        for product in delivery_products[index:index + 2]:
            lines.append(
                {
                    "sku": product["sku"],
                    "demand_qty": scale_quantity(RNG.choice(quantity_profile["quantities"]), product["uom"], "delivery"),
                    "done_qty": 0,
                    "location": delivery_location_for(product),
                }
            )
        deliveries.append(
            {
                "ref": f"DEL-{index // 2 + 1:05d}",
                "customer": customers[(index // 2) % len(customers)],
                "warehouse": "WH1",
                "scheduled_in_days": (index // 2) % 5,
                "status": ["draft", "waiting", "ready"][index % 3],
                "created_by": staff if index % 2 == 0 else manager,
                "notes": "Curated outbound demand from UCI quantity profile.",
                "lines": lines,
            }
        )

    transfers = []
    transfer_candidates = top_products(grouped["Raw Materials"], 5) + top_products(grouped["Packaging"], 4) + top_products(grouped["Electronics"], 4)
    for index, product in enumerate(transfer_candidates, start=1):
        transfers.append(
            {
                "ref": f"INT-{index:05d}",
                "from_location": primary_location_for(product),
                "to_location": transfer_destination(product),
                "scheduled_in_days": index % 4,
                "status": ["draft", "waiting", "ready"][index % 3],
                "created_by": staff if index % 2 else manager,
                "notes": "Internal replenishment between main storage and execution areas.",
                "lines": [
                    {
                        "sku": product["sku"],
                        "qty": scale_quantity(RNG.choice(quantity_profile["quantities"]), product["uom"], "transfer"),
                    }
                ],
            }
        )

    adjustments = []
    adjustment_candidates = top_products(products, 12)
    for index, product in enumerate(adjustment_candidates[:8], start=1):
        location = primary_location_for(product)
        change = -max(1, scale_quantity(RNG.choice(quantity_profile["quantities"]), product["uom"], "adjustment"))
        adjustments.append(
            {
                "ref": f"ADJ-{index:05d}",
                "location": location,
                "notes": "Cycle count correction imported from curated demand variance profile.",
                "scheduled_in_days": index % 3,
                "status": ["draft", "waiting"][index % 2],
                "created_by": manager if index % 2 else staff,
                "lines": [
                    {
                        "sku": product["sku"],
                        "change_qty": change,
                    }
                ],
            }
        )

    return {
        "receipts": receipts,
        "deliveries": deliveries,
        "transfers": transfers,
        "adjustments": adjustments,
    }


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row))
            handle.write("\n")


def main() -> None:
    ensure_dirs()
    http = session()

    all_candidates: list[dict] = []
    source_manifest = {
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "seed": SEED,
        "sources": [],
    }

    for source in AMAZON_SOURCES:
        candidates, scanned = stream_amazon_candidates(http, source)
        all_candidates.extend(candidates)
        write_jsonl(RAW_DIR / f"{source['name']}_sample.jsonl", candidates)
        source_manifest["sources"].append(
            {
                "dataset": source["dataset"],
                "source_url": source["url"],
                "records_scanned": scanned,
                "sampled_candidates": len(candidates),
            }
        )

    quantity_profile = load_uci_profile(http)
    source_manifest["sources"].append(
        {
            "dataset": quantity_profile["dataset"],
            "source_url": quantity_profile["source_url"],
            "rows_processed": quantity_profile["rows_processed"],
            "quantity_sample_size": quantity_profile["quantity_sample_size"],
            "quantity_quantiles": quantity_profile["quantity_quantiles"],
        }
    )

    selected = select_products(all_candidates)
    products = build_products(selected, quantity_profile)
    history = generate_history(products, quantity_profile)
    operations = generate_operations(products, quantity_profile)

    write_json(CURATED_DIR / "categories.json", CATEGORY_DEFINITIONS)
    write_json(CURATED_DIR / "products.json", products)
    write_json(CURATED_DIR / "operations.json", operations)
    write_json(CURATED_DIR / "history.json", history)
    write_json(CURATED_DIR / "source-manifest.json", source_manifest)

    summary = {
        "generated_at": source_manifest["generated_at"],
        "product_count": len(products),
        "history_count": len(history),
        "receipts": len(operations["receipts"]),
        "deliveries": len(operations["deliveries"]),
        "transfers": len(operations["transfers"]),
        "adjustments": len(operations["adjustments"]),
        "categories": dict(Counter(product["category"] for product in products)),
    }
    write_json(RAW_DIR / "curation_summary.json", summary)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
